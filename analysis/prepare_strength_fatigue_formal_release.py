"""Create hashed formal releases after the pre-outcome gate is verified."""
from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from audit_strength_fatigue_preoutcome import (
    PARAMETER_HEADER,
    file_digest,
    normalize_doi,
    phase_flags,
    processing_flags,
)
from common import DB, RESULTS, ensure_output_dirs
from run_state_matched_mpea_borrowing_screen import elemental_system
from scripts.localdb.build_localdb import canonical_formula


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
DESIGN = HERE / "strength_to_fatigue_ood_design.json"
AUDIT = RESULTS / "strength_fatigue_preoutcome_audit.json"
VERIFIED = RESULTS / "strength_fatigue_preoutcome_VERIFIED.json"
METADATA = RESULTS / "strength_fatigue_target_metadata_no_outcomes.csv"
XLSX = ROOT / "tmp" / "fatigue_cma2022" / "FatigueData-CMA2022.xlsx"
TARGET_RELEASE = RESULTS / "strength_fatigue_formal_target_release.csv"
DONOR_RELEASE = RESULTS / "strength_fatigue_formal_donor_release.csv"
MANIFEST = RESULTS / "strength_fatigue_formal_release_manifest.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--design", type=Path, default=DESIGN)
    parser.add_argument("--audit", type=Path, default=AUDIT)
    parser.add_argument("--verified", type=Path, default=VERIFIED)
    parser.add_argument("--metadata", type=Path, default=METADATA)
    parser.add_argument("--xlsx", type=Path, default=XLSX)
    parser.add_argument("--target-output", type=Path, default=TARGET_RELEASE)
    parser.add_argument("--donor-output", type=Path, default=DONOR_RELEASE)
    parser.add_argument("--manifest-output", type=Path, default=MANIFEST)
    return parser.parse_args()


def recipient_release(
    xlsx: Path, metadata_path: Path
) -> pd.DataFrame:
    metadata = pd.read_csv(metadata_path)
    eligible_ids = set(metadata["dataset_id"].astype(int))
    workbook = load_workbook(xlsx, read_only=True, data_only=True)

    parameter_rows = []
    for values in workbook["parameter"].iter_rows(min_row=3, values_only=True):
        if values[0] is None or int(values[0]) not in eligible_ids:
            continue
        row = dict(zip(PARAMETER_HEADER, values, strict=True))
        parameter_rows.append(
            {
                "dataset_id": int(row["dataset_id"]),
                "measured_yield_strength_mpa": pd.to_numeric(
                    row["yield_strength_mpa"], errors="coerce"
                ),
                "measured_ultimate_tensile_strength_mpa": pd.to_numeric(
                    row["ultimate_tensile_strength_mpa"], errors="coerce"
                ),
            }
        )
    parameters = pd.DataFrame(parameter_rows)

    sn_rows = []
    for values in workbook["S-N"].iter_rows(min_row=2, values_only=True):
        dataset_id, life, stress, runout = values
        if dataset_id is None or int(dataset_id) not in eligible_ids:
            continue
        sn_rows.append(
            {
                "dataset_id": int(dataset_id),
                "life_cycles": pd.to_numeric(life, errors="coerce"),
                "stress_amplitude_mpa": pd.to_numeric(stress, errors="coerce"),
                "runout": pd.to_numeric(runout, errors="coerce"),
            }
        )
    outcomes = pd.DataFrame(sn_rows)
    outcomes = outcomes[
        np.isfinite(outcomes["life_cycles"])
        & np.isfinite(outcomes["stress_amplitude_mpa"])
        & (outcomes["life_cycles"] > 0)
        & (outcomes["stress_amplitude_mpa"] > 0)
        & outcomes["runout"].isin([0, 1])
    ].copy()
    release = (
        outcomes.merge(parameters, on="dataset_id", how="left", validate="many_to_one")
        .merge(metadata, on="dataset_id", how="left", validate="many_to_one")
        .sort_values(
            [
                "provenance_chemistry_component",
                "dataset_id",
                "life_cycles",
                "stress_amplitude_mpa",
            ],
            kind="mergesort",
        )
        .reset_index(drop=True)
    )
    if release["composition_key"].isna().any():
        raise AssertionError("Formal target release lost composition metadata")
    return release


def donor_release(recipient_dois: set[str]) -> pd.DataFrame:
    with sqlite3.connect(DB) as connection:
        borg = pd.read_sql_query(
            """SELECT rowid AS raw_row_id, FORMULA,
                      [PROPERTY: UTS (MPa)] AS borg_uts,
                      [PROPERTY: YS (MPa)] AS borg_ys,
                      [PROPERTY: HV] AS borg_hardness,
                      [PROPERTY: Elongation (%)] AS borg_elongation,
                      [PROPERTY: Processing method] AS processing,
                      [PROPERTY: BCC/FCC/other] AS phase,
                      [PROPERTY: Test temperature ($^\\circ$C)] AS temperature_c,
                      [REFERENCE: doi] AS doi
               FROM raw_mpea""",
            connection,
        )
        birdshot = pd.read_sql_query(
            """SELECT source_row_id AS raw_row_id, material_key, value,
                      conditions_json, source_reference
               FROM measurements
               WHERE dataset='birdshot-high-entropy-alloy-campaign'
                 AND property='UTS_True (Mpa)'""",
            connection,
        )

    borg["doi"] = borg["doi"].map(normalize_doi)
    borg = borg[~borg["doi"].isin(recipient_dois)].copy()
    borg["material_key"] = borg["FORMULA"].map(
        lambda value: canonical_formula(value)[0]
    )
    borg = borg[borg["material_key"].notna()].copy()
    common = []
    property_map = {
        "borg_uts": "borg_uts",
        "borg_ys": "borg_ys",
        "borg_hardness": "borg_hardness",
        "borg_elongation": "borg_elongation",
    }
    for donor, column in property_map.items():
        values = pd.to_numeric(borg[column], errors="coerce")
        subset = borg[np.isfinite(values) & (values > 0)].copy()
        subset["donor"] = donor
        subset["value_raw"] = pd.to_numeric(subset[column], errors="coerce")
        subset["log_value"] = np.log10(subset["value_raw"].to_numpy(float))
        common.append(subset)
    release = pd.concat(common, ignore_index=True)
    release["elemental_system"] = release["material_key"].map(elemental_system)
    process = pd.DataFrame(
        [processing_flags(value) for value in release["processing"]]
    )
    phases = pd.DataFrame([phase_flags(value) for value in release["phase"]])
    release = pd.concat([release.reset_index(drop=True), process, phases], axis=1)
    release["temperature_c"] = pd.to_numeric(
        release["temperature_c"], errors="coerce"
    )
    release = release[
        [
            "donor",
            "raw_row_id",
            "material_key",
            "elemental_system",
            "doi",
            "value_raw",
            "log_value",
            "temperature_c",
            *process.columns,
            *phases.columns,
        ]
    ]

    birdshot["value_raw"] = pd.to_numeric(birdshot["value"], errors="coerce")
    birdshot = birdshot[
        np.isfinite(birdshot["value_raw"]) & (birdshot["value_raw"] > 0)
    ].copy()
    birdshot["donor"] = "birdshot_uts"
    birdshot["elemental_system"] = birdshot["material_key"].map(elemental_system)
    birdshot["doi"] = "10.5281/zenodo.16396374"
    birdshot["log_value"] = np.log10(birdshot["value_raw"].to_numpy(float))
    birdshot["temperature_c"] = np.nan
    for column in [*process.columns, *phases.columns]:
        birdshot[column] = 0
    birdshot = birdshot[
        [
            "donor",
            "raw_row_id",
            "material_key",
            "elemental_system",
            "doi",
            "value_raw",
            "log_value",
            "temperature_c",
            *process.columns,
            *phases.columns,
        ]
    ]
    return (
        pd.concat([release, birdshot], ignore_index=True)
        .sort_values(["donor", "material_key", "raw_row_id"], kind="mergesort")
        .reset_index(drop=True)
    )


def main() -> None:
    args = parse_args()
    ensure_output_dirs()
    design = json.loads(args.design.read_text(encoding="utf-8"))
    audit = json.loads(args.audit.read_text(encoding="utf-8"))
    verified = json.loads(args.verified.read_text(encoding="utf-8"))
    if audit["status"] != "eligible-preoutcome":
        raise AssertionError("Pre-outcome audit is not eligible")
    if verified["status"] != "verified-eligible-preoutcome":
        raise AssertionError("Pre-outcome verification is not complete")
    if audit["design_sha256"] != file_digest(args.design):
        raise AssertionError("Design changed after pre-outcome audit")
    if verified["design_sha256"] != file_digest(args.design):
        raise AssertionError("Design changed after independent verification")
    if audit["metadata_csv_sha256"] != file_digest(args.metadata):
        raise AssertionError("Outcome-free metadata changed")
    if audit["recipient_xlsx_md5"] != file_digest(args.xlsx, "md5"):
        raise AssertionError("Recipient workbook changed")

    target = recipient_release(args.xlsx, args.metadata)
    donors = donor_release(set(target["doi"].map(normalize_doi)) - {""})
    expected_source_size = design["models"]["primary_common_source_size"]
    donor_counts = donors["donor"].value_counts().to_dict()
    for donor in ("borg_uts", "borg_hardness", "borg_elongation"):
        if donor_counts.get(donor, 0) < expected_source_size:
            raise AssertionError(f"{donor} has fewer than {expected_source_size} rows")

    args.target_output.parent.mkdir(parents=True, exist_ok=True)
    target.to_csv(args.target_output, index=False)
    donors.to_csv(args.donor_output, index=False)
    manifest = {
        "status": "formal-release-after-design-freeze",
        "created_utc": datetime.now(timezone.utc).isoformat(),
        "design_sha256": file_digest(args.design),
        "preoutcome_audit_sha256": file_digest(args.audit),
        "preoutcome_verification_sha256": file_digest(args.verified),
        "target_release_sha256": file_digest(args.target_output),
        "donor_release_sha256": file_digest(args.donor_output),
        "recipient_xlsx_md5": file_digest(args.xlsx, "md5"),
        "target_rows": int(len(target)),
        "target_curves": int(target["dataset_id"].nunique()),
        "target_failure_rows": int(target["runout"].eq(0).sum()),
        "target_runout_rows": int(target["runout"].eq(1).sum()),
        "donor_rows": {key: int(value) for key, value in donor_counts.items()},
        "recipient_numeric_outcomes_released": True,
        "claim_guard": design["claim_guard"],
    }
    args.manifest_output.write_text(
        json.dumps(manifest, indent=2, sort_keys=True), encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "status": manifest["status"],
                "target_rows": manifest["target_rows"],
                "target_curves": manifest["target_curves"],
                "donor_rows": manifest["donor_rows"],
                "design_sha256": manifest["design_sha256"],
            },
            indent=2,
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
