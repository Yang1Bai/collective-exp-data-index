"""Independent semantic verifier for the strength -> fatigue pre-outcome gate."""
from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
DEFAULT_DESIGN = HERE / "strength_to_fatigue_ood_design.json"
DEFAULT_XLSX = ROOT / "tmp" / "fatigue_cma2022" / "FatigueData-CMA2022.xlsx"
DEFAULT_AUDIT = HERE / "results" / "strength_fatigue_preoutcome_audit.json"
DEFAULT_METADATA = (
    HERE / "results" / "strength_fatigue_target_metadata_no_outcomes.csv"
)
DEFAULT_OUTPUT = HERE / "results" / "strength_fatigue_preoutcome_VERIFIED.json"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--design", type=Path, default=DEFAULT_DESIGN)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--audit", type=Path, default=DEFAULT_AUDIT)
    parser.add_argument("--metadata", type=Path, default=DEFAULT_METADATA)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


def digest(path: Path, algorithm: str = "sha256") -> str:
    value = hashlib.new(algorithm)
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            value.update(block)
    return value.hexdigest()


def main() -> None:
    args = parse_args()
    design = json.loads(args.design.read_text(encoding="utf-8"))
    audit = json.loads(args.audit.read_text(encoding="utf-8"))
    metadata = pd.read_csv(args.metadata)

    if audit["status"] != "eligible-preoutcome":
        raise AssertionError("Audit is not eligible-preoutcome")
    if audit["design_sha256"] != digest(args.design):
        raise AssertionError("Design hash mismatch")
    if audit["recipient_xlsx_md5"] != digest(args.xlsx, "md5"):
        raise AssertionError("Recipient MD5 mismatch")
    if design["recipient"]["xlsx_md5"] != digest(args.xlsx, "md5"):
        raise AssertionError("Frozen recipient MD5 mismatch")
    if audit["metadata_csv_sha256"] != digest(args.metadata):
        raise AssertionError("Metadata hash mismatch")

    forbidden_columns = {
        "life",
        "cycles",
        "stress_amplitude",
        "runout",
        "yield_strength_mpa",
        "ultimate_tensile_strength_mpa",
    }
    leaked = forbidden_columns & set(metadata.columns)
    if leaked:
        raise AssertionError(f"Outcome or oracle values leaked to metadata: {leaked}")

    if len(metadata) != audit["recipient"]["eligible_curves"]:
        raise AssertionError("Eligible curve count mismatch")
    if metadata["doi"].nunique() != audit["recipient"]["unique_dois"]:
        raise AssertionError("DOI count mismatch")
    if (
        metadata["composition_key"].nunique()
        != audit["recipient"]["unique_compositions"]
    ):
        raise AssertionError("Composition count mismatch")
    if (
        metadata["provenance_chemistry_component"].nunique()
        != audit["recipient"]["provenance_chemistry_components"]
    ):
        raise AssertionError("Connected-component count mismatch")
    if not all(audit["gate_checks"].values()):
        raise AssertionError("At least one frozen pre-outcome gate did not pass")
    if audit["outcome_access"]["numeric_fatigue_outcome_cells_read"] != 0:
        raise AssertionError("Numeric fatigue outcomes were read before verification")

    workbook = load_workbook(args.xlsx, read_only=True, data_only=True)
    if set(workbook.sheetnames) != {"S-N", "e-N", "dadn", "parameter"}:
        raise AssertionError("Workbook sheet set changed")
    # Verify only the parameter row count; do not touch numeric outcome rows.
    parameter_rows = sum(
        1
        for values in workbook["parameter"].iter_rows(
            min_row=3, min_col=1, max_col=1, values_only=True
        )
        if values[0] is not None
    )
    if parameter_rows != audit["outcome_access"]["parameter_rows_read"]:
        raise AssertionError("Parameter row count mismatch")

    output = {
        "status": "verified-eligible-preoutcome",
        "design_sha256": digest(args.design),
        "audit_sha256": digest(args.audit),
        "metadata_sha256": digest(args.metadata),
        "recipient_xlsx_md5": digest(args.xlsx, "md5"),
        "eligible_curves": int(len(metadata)),
        "unique_dois": int(metadata["doi"].nunique()),
        "unique_compositions": int(metadata["composition_key"].nunique()),
        "provenance_chemistry_components": int(
            metadata["provenance_chemistry_component"].nunique()
        ),
        "numeric_fatigue_outcome_cells_read": 0,
        "claim_guard": design["claim_guard"],
    }
    args.output.write_text(
        json.dumps(output, indent=2, sort_keys=True), encoding="utf-8"
    )
    print(json.dumps(output, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
