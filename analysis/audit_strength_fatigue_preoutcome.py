"""Outcome-free feasibility audit for static-strength -> fatigue borrowing.

This script is intentionally restricted to the recipient workbook's parameter
worksheet. It must never read numeric rows from S-N, e-N, or dadn.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import re
import sqlite3
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.ensemble import ExtraTreesRegressor
from sklearn.metrics import r2_score
from sklearn.model_selection import GroupKFold

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from common import DB, ELEMENTS, RESULTS, composition_features, ensure_output_dirs
from run_state_matched_mpea_borrowing_screen import elemental_system
from scripts.localdb.build_localdb import canonical_formula


DESIGN_PATH = HERE / "strength_to_fatigue_ood_design.json"
DEFAULT_XLSX = ROOT / "tmp" / "fatigue_cma2022" / "FatigueData-CMA2022.xlsx"
AUDIT_OUTPUT = RESULTS / "strength_fatigue_preoutcome_audit.json"
METADATA_OUTPUT = RESULTS / "strength_fatigue_target_metadata_no_outcomes.csv"

PARAMETER_HEADER = [
    "dataset_id",
    "title",
    "authors",
    "publication_source",
    "publication_year",
    "institution",
    "country_region",
    "funding_agency",
    "doi",
    "fatigue_type",
    "extraction_method",
    "material_type",
    "atomic_structure",
    "material_name",
    "composition_ratio_type",
    "composition_raw",
    "glass_transition_temperature_c",
    "grain_size_um",
    "processing_raw",
    "surface_treatment",
    "ingot_shape",
    "ingot_size_mm",
    "fatigue_test_type",
    "fatigue_temperature_c",
    "fatigue_environment",
    "load_ratio",
    "frequency_hz",
    "fatigue_machine",
    "fatigue_standard",
    "load_control",
    "failure_criterion",
    "specimen_description",
    "critical_section_mm",
    "stress_concentration_factor",
    "youngs_modulus_gpa",
    "yield_strength_mpa",
    "ultimate_tensile_strength_mpa",
    "elongation_percent",
    "fracture_toughness",
    "fatigue_crack_growth_threshold",
    "rating_score",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--design", type=Path, default=DESIGN_PATH)
    parser.add_argument("--audit-output", type=Path, default=AUDIT_OUTPUT)
    parser.add_argument("--metadata-output", type=Path, default=METADATA_OUTPUT)
    return parser.parse_args()


def file_digest(path: Path, algorithm: str = "sha256") -> str:
    digest = hashlib.new(algorithm)
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalize_doi(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip().lower()
    for prefix in (
        "https://doi.org/",
        "http://doi.org/",
        "https://dx.doi.org/",
        "http://dx.doi.org/",
        "doi:",
    ):
        if text.startswith(prefix):
            text = text[len(prefix) :]
    return text.strip().rstrip(".")


def parse_composition(value: object) -> str | None:
    if value is None or pd.isna(value):
        return None
    pairs = re.findall(
        r"([A-Z][a-z]?)[^0-9.+-]*([0-9]+(?:\.[0-9]+)?)", str(value)
    )
    if len(pairs) < 2:
        return None
    formula = "".join(f"{element}{amount}" for element, amount in pairs)
    return canonical_formula(formula)[0]


def text_flag(value: object, terms: Iterable[str]) -> int:
    text = "" if value is None or pd.isna(value) else str(value).lower()
    return int(any(term in text for term in terms))


def processing_flags(value: object) -> dict[str, int]:
    return {
        "process_cast_or_melt": text_flag(
            value, ("cast", "melt", "induction", "arc ", "remelt")
        ),
        "process_wrought": text_flag(
            value, ("roll", "forge", "swag", "extrud", "wrought", "deform")
        ),
        "process_powder": text_flag(
            value, ("powder", "sinter", "mechanical alloy", "spark plasma")
        ),
        "process_heat_treated": text_flag(
            value, ("anneal", "heat treat", "homogen", "solution", "age")
        ),
        "process_additive": text_flag(
            value, ("additive", "laser", "selective", "electron beam")
        ),
    }


def phase_flags(value: object) -> dict[str, int]:
    return {
        "phase_fcc": text_flag(value, ("fcc", "face-centered", "face centred")),
        "phase_bcc": text_flag(value, ("bcc", "body-centered", "body centred")),
        "phase_hcp": text_flag(value, ("hcp", "hexagonal")),
        "phase_amorphous": text_flag(value, ("amorph", "metallic glass")),
        "phase_multiphase": text_flag(value, ("+", "dual", "multi", "mixture")),
    }


def read_parameter_metadata(xlsx: Path) -> tuple[pd.DataFrame, dict[str, Any]]:
    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    expected_sheets = {"S-N", "e-N", "dadn", "parameter"}
    if set(workbook.sheetnames) != expected_sheets:
        raise AssertionError(f"Unexpected workbook sheets: {workbook.sheetnames}")

    # Header-only inspection. No numeric cells in the outcome sheets are read.
    outcome_headers: dict[str, list[str]] = {}
    for sheet_name in ("S-N", "e-N", "dadn"):
        row = next(
            workbook[sheet_name].iter_rows(
                min_row=1, max_row=1, values_only=True
            )
        )
        outcome_headers[sheet_name] = [str(value) for value in row]

    sheet = workbook["parameter"]
    second_header = next(
        sheet.iter_rows(min_row=2, max_row=2, values_only=True)
    )
    if len(second_header) != len(PARAMETER_HEADER):
        raise AssertionError(
            f"Parameter schema has {len(second_header)} columns, expected "
            f"{len(PARAMETER_HEADER)}"
        )

    rows = [
        dict(zip(PARAMETER_HEADER, values, strict=True))
        for values in sheet.iter_rows(min_row=3, values_only=True)
        if values[0] is not None
    ]
    frame = pd.DataFrame(rows)
    frame["doi"] = frame["doi"].map(normalize_doi)
    frame["composition_key"] = frame["composition_raw"].map(parse_composition)
    for name in (
        "fatigue_type",
        "material_type",
        "fatigue_test_type",
        "fatigue_environment",
        "load_control",
    ):
        frame[name] = frame[name].fillna("").astype(str).str.strip().str.lower()
    for column in ("fatigue_temperature_c", "load_ratio", "frequency_hz"):
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    for column in ("yield_strength_mpa", "ultimate_tensile_strength_mpa"):
        frame[f"has_{column}"] = pd.to_numeric(
            frame[column], errors="coerce"
        ).notna()

    process = pd.DataFrame(
        [processing_flags(value) for value in frame["processing_raw"]]
    )
    phases = pd.DataFrame(
        [phase_flags(value) for value in frame["atomic_structure"]]
    )
    frame = pd.concat([frame, process, phases], axis=1)

    safe_columns = [
        "dataset_id",
        "title",
        "publication_year",
        "doi",
        "fatigue_type",
        "material_type",
        "atomic_structure",
        "material_name",
        "composition_raw",
        "composition_key",
        "processing_raw",
        "fatigue_test_type",
        "fatigue_temperature_c",
        "fatigue_environment",
        "load_ratio",
        "frequency_hz",
        "load_control",
        "has_yield_strength_mpa",
        "has_ultimate_tensile_strength_mpa",
        *process.columns,
        *phases.columns,
    ]
    return (
        frame[safe_columns].copy(),
        {
            "workbook_sheets": workbook.sheetnames,
            "outcome_sheet_headers": outcome_headers,
            "numeric_fatigue_outcome_cells_read": 0,
            "parameter_rows_read": int(len(frame)),
        },
    )


class UnionFind:
    def __init__(self) -> None:
        self.parent: dict[str, str] = {}

    def find(self, value: str) -> str:
        self.parent.setdefault(value, value)
        if self.parent[value] != value:
            self.parent[value] = self.find(self.parent[value])
        return self.parent[value]

    def union(self, left: str, right: str) -> None:
        root_left = self.find(left)
        root_right = self.find(right)
        if root_left != root_right:
            self.parent[root_right] = root_left


def assign_components(frame: pd.DataFrame) -> pd.Series:
    graph = UnionFind()
    for row in frame.itertuples(index=False):
        graph.union(f"doi:{row.doi}", f"composition:{row.composition_key}")
    roots = {
        node: graph.find(node)
        for node in graph.parent
    }
    ordered_roots = {
        root: index
        for index, root in enumerate(sorted(set(roots.values())))
    }
    return frame.apply(
        lambda row: (
            f"component_{ordered_roots[graph.find(f'doi:{row.doi}')]:03d}"
        ),
        axis=1,
    )


def load_borg_uts(recipient_dois: set[str]) -> pd.DataFrame:
    with sqlite3.connect(DB) as connection:
        frame = pd.read_sql_query(
            """SELECT rowid AS raw_row_id, FORMULA,
                      [PROPERTY: UTS (MPa)] AS uts_mpa,
                      [PROPERTY: Processing method] AS processing,
                      [PROPERTY: BCC/FCC/other] AS phase,
                      [PROPERTY: Test temperature ($^\\circ$C)] AS temperature_c,
                      [REFERENCE: doi] AS doi
               FROM raw_mpea""",
            connection,
        )
    frame["uts_mpa"] = pd.to_numeric(frame["uts_mpa"], errors="coerce")
    frame = frame[np.isfinite(frame["uts_mpa"]) & (frame["uts_mpa"] > 0)].copy()
    frame["doi"] = frame["doi"].map(normalize_doi)
    frame = frame[~frame["doi"].isin(recipient_dois)].copy()
    frame["composition_key"] = frame["FORMULA"].map(
        lambda value: canonical_formula(value)[0]
    )
    frame = frame[frame["composition_key"].notna()].copy()
    frame["group"] = frame["composition_key"].map(elemental_system)
    return frame.reset_index(drop=True)


def pure_composition(frame: pd.DataFrame) -> np.ndarray:
    return composition_features(frame["composition_key"].astype(str).tolist())[
        :, : len(ELEMENTS)
    ]


def nearest_source_distance(
    recipient: pd.DataFrame, source: pd.DataFrame
) -> pd.DataFrame:
    unique_target = (
        recipient[["composition_key"]]
        .drop_duplicates()
        .sort_values("composition_key")
        .reset_index(drop=True)
    )
    unique_source = (
        source[["composition_key"]]
        .drop_duplicates()
        .sort_values("composition_key")
        .reset_index(drop=True)
    )
    target_x = pure_composition(unique_target)
    source_x = pure_composition(unique_source)
    distance = np.abs(target_x[:, None, :] - source_x[None, :, :]).sum(axis=2)
    unique_target["nearest_borg_l1"] = distance.min(axis=1)
    return unique_target


def donor_group_oof_r2(frame: pd.DataFrame, seed: int, trees: int) -> float:
    comp = pure_composition(frame)
    process = pd.DataFrame(
        [processing_flags(value) for value in frame["processing"]]
    ).to_numpy(float)
    phase = pd.DataFrame(
        [phase_flags(value) for value in frame["phase"]]
    ).to_numpy(float)
    temperature = (
        pd.to_numeric(frame["temperature_c"], errors="coerce")
        .fillna(pd.to_numeric(frame["temperature_c"], errors="coerce").median())
        .to_numpy(float)[:, None]
    )
    x = np.concatenate([comp, process, phase, temperature], axis=1)
    y = np.log10(frame["uts_mpa"].to_numpy(float))
    groups = frame["group"].astype(str).to_numpy()
    folds = min(5, len(np.unique(groups)))
    if folds < 2:
        return math.nan
    prediction = np.full(len(frame), np.nan)
    splitter = GroupKFold(n_splits=folds)
    for fold, (fit, held) in enumerate(splitter.split(x, y, groups)):
        model = ExtraTreesRegressor(
            n_estimators=320,
            min_samples_leaf=2,
            max_features=0.7,
            random_state=seed + fold,
            n_jobs=-1,
        )
        model.fit(x[fit], y[fit])
        prediction[held] = model.predict(x[held])
    if not np.isfinite(prediction).all():
        raise AssertionError("Incomplete Borg group-OOF prediction")
    return float(r2_score(y, prediction))


def main() -> None:
    args = parse_args()
    ensure_output_dirs()
    design = json.loads(args.design.read_text(encoding="utf-8"))
    design_sha256 = file_digest(args.design)
    if file_digest(args.xlsx, "md5") != design["recipient"]["xlsx_md5"]:
        raise AssertionError("Recipient XLSX MD5 changed")

    all_metadata, access_audit = read_parameter_metadata(args.xlsx)
    recipient = all_metadata[
        all_metadata["material_type"].eq(
            design["recipient"]["eligible_material_type"]
        )
        & all_metadata["fatigue_type"].eq(
            design["recipient"]["eligible_fatigue_type"]
        )
    ].copy()
    recipient["provenance_chemistry_component"] = assign_components(recipient)

    recipient_dois = set(recipient["doi"]) - {""}
    borg = load_borg_uts(recipient_dois)
    distances = nearest_source_distance(recipient, borg)
    recipient = recipient.merge(distances, on="composition_key", how="left")

    gates_spec = design["preoutcome_gate"]
    unique_compositions = int(recipient["composition_key"].nunique())
    supported_compositions = int(
        distances["nearest_borg_l1"]
        .le(gates_spec["supported_neighbor_l1_threshold"])
        .sum()
    )
    parseable_fraction = float(recipient["composition_key"].notna().mean())
    temperature_fraction = float(recipient["fatigue_temperature_c"].notna().mean())
    ratio_fraction = float(recipient["load_ratio"].notna().mean())
    processing_fraction = float(
        recipient["processing_raw"].fillna("").astype(str).str.strip().ne("").mean()
    )
    source_oof_r2 = donor_group_oof_r2(
        borg, seed=design["models"]["random_seed"], trees=320
    )
    checks = {
        "eligible_curves": len(recipient) >= gates_spec["minimum_eligible_curves"],
        "unique_recipient_dois": (
            len(recipient_dois) >= gates_spec["minimum_unique_recipient_dois"]
        ),
        "unique_recipient_compositions": (
            unique_compositions
            >= gates_spec["minimum_unique_recipient_compositions"]
        ),
        "parseable_composition_fraction": (
            parseable_fraction
            >= gates_spec["minimum_parseable_composition_fraction"]
        ),
        "test_temperature_fraction": (
            temperature_fraction
            >= gates_spec["minimum_curves_with_test_temperature_fraction"]
        ),
        "load_ratio_fraction": (
            ratio_fraction
            >= gates_spec["minimum_curves_with_load_ratio_fraction"]
        ),
        "processing_fraction": (
            processing_fraction
            >= gates_spec["minimum_curves_with_processing_fraction"]
        ),
        "supported_recipient_compositions": (
            supported_compositions
            >= gates_spec[
                "minimum_recipient_compositions_with_supported_borg_neighbor"
            ]
        ),
        "nonoverlapping_borg_uts_rows": (
            len(borg) >= gates_spec["minimum_nonoverlapping_borg_uts_rows"]
        ),
        "source_group_oof_r2": (
            source_oof_r2 >= gates_spec["minimum_source_group_oof_r2"]
        ),
        "numeric_fatigue_outcomes_unread": (
            access_audit["numeric_fatigue_outcome_cells_read"] == 0
        ),
    }
    status = "eligible-preoutcome" if all(checks.values()) else "ineligible-preoutcome"

    recipient = recipient.sort_values(
        ["provenance_chemistry_component", "doi", "dataset_id"],
        kind="mergesort",
    ).reset_index(drop=True)
    args.metadata_output.parent.mkdir(parents=True, exist_ok=True)
    recipient.to_csv(args.metadata_output, index=False)

    overlap_dois = set(
        pd.read_sql_query(
            "SELECT [REFERENCE: doi] AS doi FROM raw_mpea",
            sqlite3.connect(DB),
        )["doi"].map(normalize_doi)
    ) & recipient_dois
    audit = {
        "status": status,
        "created_utc": datetime.now(timezone.utc).isoformat(),
        "design_sha256": design_sha256,
        "recipient_xlsx_sha256": file_digest(args.xlsx),
        "recipient_xlsx_md5": file_digest(args.xlsx, "md5"),
        "metadata_csv_sha256": file_digest(args.metadata_output),
        "outcome_access": access_audit,
        "recipient": {
            "eligible_curves": int(len(recipient)),
            "unique_dois": int(len(recipient_dois)),
            "unique_compositions": unique_compositions,
            "provenance_chemistry_components": int(
                recipient["provenance_chemistry_component"].nunique()
            ),
            "parseable_composition_fraction": parseable_fraction,
            "temperature_present_fraction": temperature_fraction,
            "load_ratio_present_fraction": ratio_fraction,
            "processing_present_fraction": processing_fraction,
            "measured_uts_available_curves": int(
                recipient["has_ultimate_tensile_strength_mpa"].sum()
            ),
            "measured_ys_available_curves": int(
                recipient["has_yield_strength_mpa"].sum()
            ),
            "test_type_counts": dict(
                Counter(recipient["fatigue_test_type"].astype(str))
            ),
        },
        "provenance": {
            "borg_recipient_doi_overlap_before_exclusion": int(len(overlap_dois)),
            "overlap_dois": sorted(overlap_dois),
            "borg_uts_rows_after_all_recipient_doi_exclusion": int(len(borg)),
            "borg_uts_groups_after_exclusion": int(borg["group"].nunique()),
        },
        "applicability": {
            "supported_neighbor_l1_threshold": gates_spec[
                "supported_neighbor_l1_threshold"
            ],
            "supported_unique_compositions": supported_compositions,
            "exact_composition_overlap": int(
                distances["nearest_borg_l1"].le(1e-12).sum()
            ),
            "nearest_borg_l1_quantiles": {
                str(quantile): float(
                    distances["nearest_borg_l1"].quantile(quantile)
                )
                for quantile in (0.0, 0.25, 0.5, 0.75, 1.0)
            },
        },
        "source_skill": {
            "borg_group_oof_r2_log10_uts": source_oof_r2,
            "outcome": "log10 UTS MPa",
            "group": "unordered elemental system",
        },
        "gate_checks": checks,
        "claim_guard": design["claim_guard"],
    }
    args.audit_output.write_text(
        json.dumps(audit, indent=2, sort_keys=True), encoding="utf-8"
    )
    print(json.dumps(audit, indent=2, sort_keys=True))
    if status != "eligible-preoutcome":
        raise SystemExit("Pre-outcome fatigue borrowing gate failed")


if __name__ == "__main__":
    main()
